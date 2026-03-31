#!/usr/bin/env python3
"""Generate a formatted Excel report from CSV input to XLSX output."""

import sys, os, re

missing = []
for pkg, imp in [("openpyxl","openpyxl"),("pyyaml","yaml")]:
    try: __import__(imp)
    except ImportError: missing.append(pkg)
if missing:
    print("Missing packages: " + " ".join(missing))
    print("Fix: pip install " + " ".join(missing))
    sys.exit(1)

import csv
import yaml
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

# ── helpers ──────────────────────────────────────────────────────────────────
def bd():
    s = Side(style='thin', color='BDBDBD')
    return Border(top=s, bottom=s, left=s, right=s)

def fill(c): return PatternFill("solid", fgColor=c)
def hfont(color='FFFFFF', size=10): return Font(name='Arial', bold=True, color=color, size=size)
def cfont(bold=False, size=10): return Font(name='Arial', bold=bold, size=size)
def ctr(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def lft(): return Alignment(horizontal='left', vertical='center', wrap_text=True)
def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w

# ── read inputs ───────────────────────────────────────────────────────────────
csv_path = sys.argv[1] if len(sys.argv) >= 2 else 'data/output/result.csv'
out = sys.argv[2] if len(sys.argv) >= 3 else 'data/output/result.xlsx'
people_path = sys.argv[3] if len(sys.argv) >= 4 else 'data/input/people.csv'
cfg_path = 'data/input/config.yaml'

if not os.path.exists(csv_path):
    print(f"Error: {csv_path} not found — run cargo first"); sys.exit(1)
if not os.path.exists(people_path):
    print(f"Error: people file not found: {people_path}"); sys.exit(1)

cfg = yaml.safe_load(open(cfg_path))
dessert_addr = f"{cfg['dessert_address']} {cfg['dessert_postal_code']} {cfg['dessert_city']}"
event_title = cfg.get('event_title', 'Happy agape 2')

rows = []
with open(csv_path) as f:
    for r in csv.DictReader(f):
        rows.append(r)

# ── load distance cache for walk times ───────────────────────────────────────
import json
cache_path = 'data/cache/distance_cache.json'
dist = {}
if os.path.exists(cache_path):
    dist = json.load(open(cache_path)).get('entries', {})

geocode_cache_path = 'data/cache/geocode_cache.json'
geocode_cache = {}
if os.path.exists(geocode_cache_path):
    raw_geo = json.load(open(geocode_cache_path))
    if isinstance(raw_geo, dict):
        geocode_cache = raw_geo
geocode_disabled = False

def normalize_address(address):
    s = (address or '').lower()
    s = s.replace('-', ' ').replace('_', ' ')
    s = re.sub(r'[^\w\s]', ' ', s, flags=re.UNICODE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def canonical_cache_key(addr_a, addr_b):
    a = normalize_address(addr_a)
    b = normalize_address(addr_b)
    return f"{a}|||{b}" if a <= b else f"{b}|||{a}"

def coord_key(lat, lon):
    return f"{float(lat):.7f},{float(lon):.7f}"

def canonical_coord_cache_key(latlon_a, latlon_b):
    a = coord_key(latlon_a[0], latlon_a[1])
    b = coord_key(latlon_b[0], latlon_b[1])
    return f"{a}|||{b}" if a <= b else f"{b}|||{a}"

# Build normalization-insensitive index from existing cache entries.
dist_normalized = {}
dist_coords = {}
for k, v in dist.items():
    if '|||' not in k:
        continue
    left, right = k.split('|||', 1)
    coordish = (
        re.fullmatch(r'\s*-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?\s*', left) and
        re.fullmatch(r'\s*-?\d+(?:\.\d+)?\s*,\s*-?\d+(?:\.\d+)?\s*', right)
    )
    if coordish:
        llat, llon = [float(x.strip()) for x in left.split(',', 1)]
        rlat, rlon = [float(x.strip()) for x in right.split(',', 1)]
        ck = canonical_coord_cache_key((llat, llon), (rlat, rlon))
        if ck not in dist_coords or v < dist_coords[ck]:
            dist_coords[ck] = v
        continue
    ck = canonical_cache_key(left, right)
    if ck not in dist_normalized or v < dist_normalized[ck]:
        dist_normalized[ck] = v

def walk(addr_from, addr_to):
    """Return walk time in minutes from cache, or 0 if not found."""
    if normalize_address(addr_from) == normalize_address(addr_to):
        return 0.0

    latlon_from = geocode_address(addr_from, cfg)
    latlon_to = geocode_address(addr_to, cfg)
    if latlon_from is not None and latlon_to is not None:
        key = canonical_coord_cache_key(latlon_from, latlon_to)
        if key in dist_coords:
            return round(dist_coords[key] / 60.0, 1)

    key = canonical_cache_key(addr_from, addr_to)
    if key in dist_normalized:
        return round(dist_normalized[key] / 60.0, 1)

    # Walking time is assumed symmetric: A->B == B->A.
    key = f"{addr_from}|||{addr_to}" if addr_from <= addr_to else f"{addr_to}|||{addr_from}"
    if key in dist:
        return round(dist[key] / 60.0, 1)
    # Backward compatibility with older directional cache files.
    legacy = f"{addr_from}|||{addr_to}"
    reverse = f"{addr_to}|||{addr_from}"
    return round(dist.get(legacy, dist.get(reverse, 0)) / 60.0, 1)

def geocode_with_google(address, api_key):
    import urllib.parse, urllib.request
    url = (
        "https://maps.googleapis.com/maps/api/geocode/json"
        f"?address={urllib.parse.quote(address)}"
        f"&key={urllib.parse.quote(api_key)}"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "shared-dinners-distributor/1.0"})
    with urllib.request.urlopen(req, timeout=2.5) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    if (payload.get("status") or "") != "OK":
        return None
    results = payload.get("results") or []
    if not results:
        return None
    location = (((results[0] or {}).get("geometry") or {}).get("location")) or {}
    lat = location.get("lat")
    lon = location.get("lng")
    if lat is None or lon is None:
        return None
    return (float(lat), float(lon))

def geocode_with_nominatim(address):
    import urllib.parse, urllib.request
    url = (
        "https://nominatim.openstreetmap.org/search"
        f"?q={urllib.parse.quote(address)}&format=json&limit=1"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "shared-dinners-distributor/1.0"})
    with urllib.request.urlopen(req, timeout=2.5) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    if not payload:
        return None
    lat = float(payload[0]["lat"])
    lon = float(payload[0]["lon"])
    return (lat, lon)

def geocode_address(address, cfg):
    global geocode_disabled
    if geocode_disabled:
        return None
    key = normalize_address(address)
    cached = geocode_cache.get(key)
    if isinstance(cached, dict) and "lat" in cached and "lon" in cached:
        return (float(cached["lat"]), float(cached["lon"]))
    if isinstance(cached, list) and len(cached) == 2:
        return (float(cached[0]), float(cached[1]))

    latlon = None
    google_failed = False
    nom_failed = False
    api_key = str(cfg.get("google_maps_api_key", "") or "").strip()
    if api_key and api_key != "YOUR_GOOGLE_MAPS_API_KEY_HERE":
        try:
            latlon = geocode_with_google(address, api_key)
        except Exception:
            google_failed = True
            latlon = None
    if latlon is None:
        try:
            latlon = geocode_with_nominatim(address)
        except Exception:
            nom_failed = True
            latlon = None
    if latlon is None and (google_failed or not api_key) and nom_failed:
        geocode_disabled = True

    if latlon is not None:
        geocode_cache[key] = {"lat": latlon[0], "lon": latlon[1]}
    return latlon

def yn(v):
    return str(v or '').strip().lower() in {'yes', 'y', 'true', '1', 'oui'}

def pick(r, *keys, default=''):
    for k in keys:
        if k in r and str(r[k]).strip() != '':
            return r[k]
    return default

def parse_int(v, default=None):
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return default

def xml_esc(v):
    import html
    return html.escape(str(v or ''), quote=True)

def write_kml_placemark(kf, name, address, style_id, category_label, lat=None, lon=None):
    kf.write("    <Placemark>\n")
    kf.write(f"      <name>{xml_esc(name)}</name>\n")
    kf.write(f"      <styleUrl>#{style_id}</styleUrl>\n")
    desc = f"Categorie: {category_label}<br/>Adresse: {xml_esc(address)}"
    kf.write(f"      <description><![CDATA[{desc}]]></description>\n")
    if address:
        kf.write(f"      <address>{xml_esc(address)}</address>\n")
    if lat is not None and lon is not None:
        kf.write("      <Point>\n")
        kf.write(f"        <coordinates>{lon:.6f},{lat:.6f},0</coordinates>\n")
        kf.write("      </Point>\n")
    kf.write("    </Placemark>\n")

def write_had_styles(kf):
    # Match icon/colors/style maps from data/output/had.kml
    kf.write("""    <Style id="icon-1517-0288D1-normal">
      <IconStyle>
        <color>ffd18802</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>0</scale>
      </LabelStyle>
    </Style>
    <Style id="icon-1517-0288D1-highlight">
      <IconStyle>
        <color>ffd18802</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>1</scale>
      </LabelStyle>
    </Style>
    <StyleMap id="icon-1517-0288D1">
      <Pair>
        <key>normal</key>
        <styleUrl>#icon-1517-0288D1-normal</styleUrl>
      </Pair>
      <Pair>
        <key>highlight</key>
        <styleUrl>#icon-1517-0288D1-highlight</styleUrl>
      </Pair>
    </StyleMap>
    <Style id="icon-1603-0288D1-normal">
      <IconStyle>
        <color>ffd18802</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>0</scale>
      </LabelStyle>
    </Style>
    <Style id="icon-1603-0288D1-highlight">
      <IconStyle>
        <color>ffd18802</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>1</scale>
      </LabelStyle>
    </Style>
    <StyleMap id="icon-1603-0288D1">
      <Pair>
        <key>normal</key>
        <styleUrl>#icon-1603-0288D1-normal</styleUrl>
      </Pair>
      <Pair>
        <key>highlight</key>
        <styleUrl>#icon-1603-0288D1-highlight</styleUrl>
      </Pair>
    </StyleMap>
    <Style id="icon-1517-BDBDBD-normal">
      <IconStyle>
        <color>ffbdbdbd</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>0</scale>
      </LabelStyle>
    </Style>
    <Style id="icon-1517-BDBDBD-highlight">
      <IconStyle>
        <color>ffbdbdbd</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>1</scale>
      </LabelStyle>
    </Style>
    <StyleMap id="icon-1517-BDBDBD">
      <Pair>
        <key>normal</key>
        <styleUrl>#icon-1517-BDBDBD-normal</styleUrl>
      </Pair>
      <Pair>
        <key>highlight</key>
        <styleUrl>#icon-1517-BDBDBD-highlight</styleUrl>
      </Pair>
    </StyleMap>
    <Style id="icon-1577-BDBDBD-normal">
      <IconStyle>
        <color>ffbdbdbd</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>0</scale>
      </LabelStyle>
    </Style>
    <Style id="icon-1577-BDBDBD-highlight">
      <IconStyle>
        <color>ffbdbdbd</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>1</scale>
      </LabelStyle>
    </Style>
    <StyleMap id="icon-1577-BDBDBD">
      <Pair>
        <key>normal</key>
        <styleUrl>#icon-1577-BDBDBD-normal</styleUrl>
      </Pair>
      <Pair>
        <key>highlight</key>
        <styleUrl>#icon-1577-BDBDBD-highlight</styleUrl>
      </Pair>
    </StyleMap>
    <Style id="icon-1577-FF5252-normal">
      <IconStyle>
        <color>ff5252ff</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>0</scale>
      </LabelStyle>
    </Style>
    <Style id="icon-1577-FF5252-highlight">
      <IconStyle>
        <color>ff5252ff</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>1</scale>
      </LabelStyle>
    </Style>
    <StyleMap id="icon-1577-FF5252">
      <Pair>
        <key>normal</key>
        <styleUrl>#icon-1577-FF5252-normal</styleUrl>
      </Pair>
      <Pair>
        <key>highlight</key>
        <styleUrl>#icon-1577-FF5252-highlight</styleUrl>
      </Pair>
    </StyleMap>
    <Style id="icon-1762-7CB342-nodesc-normal">
      <IconStyle>
        <color>ff42b37c</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>0</scale>
      </LabelStyle>
      <BalloonStyle>
        <text><![CDATA[<h3>$[name]</h3>]]></text>
      </BalloonStyle>
    </Style>
    <Style id="icon-1762-7CB342-nodesc-highlight">
      <IconStyle>
        <color>ff42b37c</color>
        <scale>1</scale>
        <Icon>
          <href>https://www.gstatic.com/mapspro/images/stock/503-wht-blank_maps.png</href>
        </Icon>
      </IconStyle>
      <LabelStyle>
        <scale>1</scale>
      </LabelStyle>
      <BalloonStyle>
        <text><![CDATA[<h3>$[name]</h3>]]></text>
      </BalloonStyle>
    </Style>
    <StyleMap id="icon-1762-7CB342-nodesc">
      <Pair>
        <key>normal</key>
        <styleUrl>#icon-1762-7CB342-nodesc-normal</styleUrl>
      </Pair>
      <Pair>
        <key>highlight</key>
        <styleUrl>#icon-1762-7CB342-nodesc-highlight</styleUrl>
      </Pair>
    </StyleMap>
""")

# Build address map: name -> address
people_csv = []
current_year = date.today().year
with open(people_path) as f:
    for r in csv.DictReader(f):
        addr = f"{pick(r, 'postal_address')} {pick(r, 'postal_code')} {pick(r, 'city')}".strip()
        can_drinks = yn(pick(r, 'recieving_for_drinks', 'receiving_for_drinks'))
        can_dinner = yn(pick(r, 'recieving_for_dinner', 'receiving_for_dinner'))
        yob = parse_int(pick(r, 'year_of_birth'), default=None)
        people_csv.append({
            'name': pick(r, 'name'),
            'gender': pick(r, 'gender', default=''),
            'year_of_birth': yob,
            'age': (current_year - yob) if yob is not None else None,
            'addr': addr,
            'can_drinks': can_drinks,
            'can_dinner': can_dinner,
            'max_drinks': parse_int(pick(r, 'number_max_recieving_for_drinks', 'number_max_receiving_for_drinks', default='0'), default=0),
            'max_dinner': parse_int(pick(r, 'number_max_recieving_for_dinner', 'number_max_receiving_for_dinner', default='0'), default=0),
        })

addr_map = {p['name']: p['addr'] for p in people_csv}
person_info = {p['name']: p for p in people_csv}

def display_person(name):
    info = person_info.get(name, {})
    yob = info.get('year_of_birth')
    return f"{name} ({yob})" if yob is not None else name

def group_stats_text(guest_rows):
    ages = []
    m_count = 0
    f_count = 0
    for g in guest_rows:
        info = person_info.get(g['name'])
        if not info:
            continue
        if info.get('age') is not None:
            ages.append(info['age'])
        gender = str(info.get('gender', '')).strip().lower()
        if gender in {'m', 'male', 'homme', 'garcon', 'garçon'}:
            m_count += 1
        elif gender in {'f', 'female', 'femme', 'fille'}:
            f_count += 1

    age_txt = "Age min/moy/max: n/d"
    if ages:
        avg_age = sum(ages) / len(ages)
        age_txt = f"Age min/moy/max: {min(ages)}/{avg_age:.1f}/{max(ages)} ans"

    mf_total = m_count + f_count
    mix_txt = "Mixite: n/d"
    if mf_total > 0:
        m_pct = 100.0 * m_count / mf_total
        f_pct = 100.0 * f_count / mf_total
        mix_txt = f"Mixite: {m_pct:.0f}% garcons / {f_pct:.0f}% filles"

    return f"{age_txt}  |  {mix_txt}"

# Compute walk times for each person
for r in rows:
    person_addr      = addr_map.get(r['name'], '')
    drinks_host_addr = addr_map.get(r['drinks_host'], '')
    dinner_host_addr = addr_map.get(r['dinner_host'], '')
    r['w1'] = walk(person_addr, drinks_host_addr)
    r['w2'] = walk(drinks_host_addr, dinner_host_addr)
    r['w3'] = walk(dinner_host_addr, dessert_addr)
    r['total'] = round(r['w1'] + r['w2'] + r['w3'], 1)

# ── build workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ═══ SHEET 1 — Overview ══════════════════════════════════════════════════════
ws = wb.active
ws.title = "Vue d'ensemble"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

ws.merge_cells("A1:G1")
ws["A1"] = f"🎉 {event_title.upper()} — SUGGESTION"
ws["A1"].font = Font(name='Arial', bold=True, size=17, color='FFFFFF')
ws["A1"].fill = fill("0B132B"); ws["A1"].alignment = ctr()
ws.row_dimensions[1].height = 40

hdrs   = ["👤 Nom","🍸 Aperitif chez","🍽️ Diner chez","🚶 Maison→Apero (min)","🚶 Apero→Diner (min)","🚶 Diner→Dessert (min)","📊 Total marche (min)"]
colors = ["1D4ED8","1D4ED8","C2410C","059669","059669","7C3AED","B91C1C"]
widths = [48,40,40,18,18,20,17]
ws.row_dimensions[2].height = 26
for ci,(h,c,w) in enumerate(zip(hdrs,colors,widths),1):
    cell = ws.cell(row=2,column=ci,value=h)
    cell.font=hfont(); cell.fill=fill(c); cell.alignment=ctr(); cell.border=bd()
    cw(ws,ci,w)

for ri,r in enumerate(rows,3):
    ws.row_dimensions[ri].height = 19
    same = r['drinks_host']==r['dinner_host']
    bg = fill("FFCCBC") if same else fill("F5F5F5" if ri%2==0 else "FFFFFF")
    vals = [display_person(r['name']),display_person(r['drinks_host']),display_person(r['dinner_host']),
            r['w1'],r['w2'],r['w3'],f"=D{ri}+E{ri}+F{ri}"]
    for ci,v in enumerate(vals,1):
        c = ws.cell(row=ri,column=ci,value=v)
        c.fill=bg; c.font=cfont(bold=(ci==1))
        c.alignment=lft() if ci <= 3 else ctr(); c.border=bd()
        if ci>=4: c.number_format='0.0'

avg_row = len(rows)+3
ws.row_dimensions[avg_row].height = 22
ws.merge_cells(f"A{avg_row}:C{avg_row}")
ws[f"A{avg_row}"]="MOYENNE"; ws[f"A{avg_row}"].font=hfont(color='000000')
ws[f"A{avg_row}"].fill=fill("E8F5E9"); ws[f"A{avg_row}"].alignment=ctr(); ws[f"A{avg_row}"].border=bd()
for ci in range(4,8):
    col=get_column_letter(ci)
    c=ws.cell(row=avg_row,column=ci,value=f"=AVERAGE({col}3:{col}{len(rows)+2})")
    c.font=hfont(color='000000',size=10); c.fill=fill("E8F5E9")
    c.number_format='0.0'; c.alignment=ctr(); c.border=bd()

max_row = len(rows)+4
ws.row_dimensions[max_row].height = 22
ws.merge_cells(f"A{max_row}:C{max_row}")
ws[f"A{max_row}"]="MAXIMUM"; ws[f"A{max_row}"].font=hfont(color='000000')
ws[f"A{max_row}"].fill=fill("E3F2FD"); ws[f"A{max_row}"].alignment=ctr(); ws[f"A{max_row}"].border=bd()
for ci in range(4,8):
    col=get_column_letter(ci)
    c=ws.cell(row=max_row,column=ci,value=f"=MAX({col}3:{col}{len(rows)+2})")
    c.font=hfont(color='000000',size=10); c.fill=fill("E3F2FD")
    c.number_format='0.0'; c.alignment=ctr(); c.border=bd()

ws.conditional_formatting.add(f"G3:G{len(rows)+2}",
    DataBarRule(start_type='min',end_type='max',color="0284C7",showValue=True))

# ═══ SHEET 2 — Aperitif ══════════════════════════════════════════════════════
from collections import defaultdict
ws2 = wb.create_sheet("Aperitif")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:D1")
ws2["A1"]="🍸 APERITIF — Repartition par hote"
ws2["A1"].font=Font(name='Arial',bold=True,size=14,color='FFFFFF')
ws2["A1"].fill=fill("1565C0"); ws2["A1"].alignment=ctr()
ws2.row_dimensions[1].height=30
for col,w in zip([1,2,3,4],[46,20,42,20]): cw(ws2,col,w)

drinks_groups = defaultdict(list)
for r in rows:
    host = r['drinks_host']
    drinks_groups.setdefault(host, [])
    drinks_groups[host].append(r)

row=2
for host,guests in drinks_groups.items():
    host_addr = addr_map.get(host,'')
    ws2.row_dimensions[row].height=26
    ws2.merge_cells(f"A{row}:D{row}")
    ws2[f"A{row}"]=f"Chez {display_person(host)}  —  {host_addr}"
    ws2[f"A{row}"].font=Font(name='Arial',bold=True,size=11,color='FFFFFF')
    ws2[f"A{row}"].fill=fill("0D47A1"); ws2[f"A{row}"].alignment=lft(); row+=1
    ws2.row_dimensions[row].height=18
    ws2.merge_cells(f"A{row}:D{row}")
    ws2[f"A{row}"]=group_stats_text(guests)
    ws2[f"A{row}"].font=Font(name='Arial',bold=True,size=9,color='0D47A1')
    ws2[f"A{row}"].fill=fill("BBDEFB"); ws2[f"A{row}"].alignment=lft(); ws2[f"A{row}"].border=bd(); row+=1
    ws2.row_dimensions[row].height=20
    for ci,h in enumerate(["👥 Participant","🚶 Maison → ici (min)","🍽️ Diner chez","🚶 Ici → Diner (min)"],1):
        c=ws2.cell(row=row,column=ci,value=h)
        c.font=hfont(size=9); c.fill=fill("42A5F5"); c.alignment=ctr(); c.border=bd()
    row+=1
    if not guests:
        ws2.row_dimensions[row].height=18
        for ci,v in enumerate(["(aucun participant)","","",""],1):
            c=ws2.cell(row=row,column=ci,value=v)
            c.fill=fill("E3F2FD"); c.font=cfont(bold=(ci==1))
            c.alignment=lft() if ci==1 else ctr(); c.border=bd()
        row+=1
    else:
        for g in guests:
            ws2.row_dimensions[row].height=18
            for ci,v in enumerate([display_person(g['name']),g['w1'],display_person(g['dinner_host']),g['w2']],1):
                c=ws2.cell(row=row,column=ci,value=v)
                c.fill=fill("E3F2FD"); c.font=cfont(bold=(ci==1))
                c.alignment=lft() if ci in (1,3) else ctr(); c.border=bd()
                if ci in (2,4): c.number_format='0.0'
            row+=1
    row+=1

# ═══ SHEET 3 — Diner ═════════════════════════════════════════════════════════
ws3=wb.create_sheet("Diner")
ws3.sheet_view.showGridLines=False
ws3.merge_cells("A1:D1")
ws3["A1"]="🍽️ DINER — Repartition par hote"
ws3["A1"].font=Font(name='Arial',bold=True,size=14,color='FFFFFF')
ws3["A1"].fill=fill("E65100"); ws3["A1"].alignment=ctr()
ws3.row_dimensions[1].height=30
for col,w in zip([1,2,3,4],[46,22,40,20]): cw(ws3,col,w)

dinner_groups = defaultdict(list)
for r in rows:
    host = r['dinner_host']
    dinner_groups.setdefault(host, [])
    dinner_groups[host].append(r)

row=2
for host,guests in dinner_groups.items():
    host_addr=addr_map.get(host,'')
    ws3.row_dimensions[row].height=26
    ws3.merge_cells(f"A{row}:D{row}")
    ws3[f"A{row}"]=f"Chez {display_person(host)}  —  {host_addr}"
    ws3[f"A{row}"].font=Font(name='Arial',bold=True,size=11,color='FFFFFF')
    ws3[f"A{row}"].fill=fill("BF360C"); ws3[f"A{row}"].alignment=lft(); row+=1
    ws3.row_dimensions[row].height=18
    ws3.merge_cells(f"A{row}:D{row}")
    ws3[f"A{row}"]=group_stats_text(guests)
    ws3[f"A{row}"].font=Font(name='Arial',bold=True,size=9,color='BF360C')
    ws3[f"A{row}"].fill=fill("FFE0B2"); ws3[f"A{row}"].alignment=lft(); ws3[f"A{row}"].border=bd(); row+=1
    ws3.row_dimensions[row].height=20
    for ci,h in enumerate(["👥 Participant","🚶 Apero → ici (min)","🍸 Apero chez","🚶 Ici → Dessert (min)"],1):
        c=ws3.cell(row=row,column=ci,value=h)
        c.font=hfont(size=9); c.fill=fill("FF7043"); c.alignment=ctr(); c.border=bd()
    row+=1
    if not guests:
        ws3.row_dimensions[row].height=18
        for ci,v in enumerate(["(aucun participant)","","",""],1):
            c=ws3.cell(row=row,column=ci,value=v)
            c.fill=fill("FFF3E0"); c.font=cfont(bold=(ci==1))
            c.alignment=lft() if ci==1 else ctr(); c.border=bd()
        row+=1
    else:
        for g in guests:
            ws3.row_dimensions[row].height=18
            for ci,v in enumerate([display_person(g['name']),g['w2'],display_person(g['drinks_host']),g['w3']],1):
                c=ws3.cell(row=row,column=ci,value=v)
                c.fill=fill("FFF3E0")
                c.font=cfont(bold=(ci==1)); c.alignment=lft() if ci in (1,3) else ctr(); c.border=bd()
                if ci in (2,4): c.number_format='0.0'
            row+=1
    row+=1

# ═══ SHEET 4 — Hosts potential vs actual ═════════════════════════════════════
ws4=wb.create_sheet("Hotes potentiels")
ws4.sheet_view.showGridLines=False
ws4.freeze_panes="A3"
for col,w in zip([1,2,3,4,5,6,7,8],[30,14,16,14,16,12,12,40]): cw(ws4,col,w)
ws4.merge_cells("A1:H1")
ws4["A1"]="🏠 HOTES POTENTIELS VS REELS"
ws4["A1"].font=Font(name='Arial',bold=True,size=14,color='FFFFFF')
ws4["A1"].fill=fill("1B5E20"); ws4["A1"].alignment=ctr(); ws4.row_dimensions[1].height=30

headers = [
    "👤 Nom",
    "🍸 Apero possible",
    "✅ A recu apero",
    "🍽️ Diner possible",
    "✅ A recu diner",
    "📦 Cap. apero",
    "📦 Cap. diner",
    "📍 Adresse",
]
ws4.row_dimensions[2].height=24
for ci,h in enumerate(headers,1):
    c=ws4.cell(row=2,column=ci,value=h)
    c.font=hfont(size=9); c.fill=fill("2E7D32"); c.alignment=ctr(); c.border=bd()

hosted_drinks = set(drinks_groups.keys())
hosted_dinner = set(dinner_groups.keys())
potential_hosts = [p for p in people_csv if p['can_drinks'] or p['can_dinner']]
potential_hosts.sort(key=lambda p: p['name'].casefold())

if not potential_hosts:
    ws4.row_dimensions[3].height=20
    ws4.merge_cells("A3:H3")
    c = ws4["A3"]
    c.value = "(aucun hote potentiel dans le fichier participants)"
    c.font = cfont(bold=True, size=10)
    c.fill = fill("E8F5E9")
    c.alignment = ctr()
    c.border = bd()
else:
    for ri,p in enumerate(potential_hosts,3):
        ws4.row_dimensions[ri].height=19
        bg = fill("E8F5E9" if ri%2==0 else "FFFFFF")
        apero_real = "-" if not p['can_drinks'] else ("Oui" if p['name'] in hosted_drinks else "Non")
        diner_real = "-" if not p['can_dinner'] else ("Oui" if p['name'] in hosted_dinner else "Non")
        vals = [
            p['name'],
            "Oui" if p['can_drinks'] else "Non",
            apero_real,
            "Oui" if p['can_dinner'] else "Non",
            diner_real,
            p['max_drinks'] if p['can_drinks'] else "-",
            p['max_dinner'] if p['can_dinner'] else "-",
            p['addr'],
        ]
        for ci,v in enumerate(vals,1):
            c=ws4.cell(row=ri,column=ci,value=v)
            c.fill=bg
            c.font=cfont(bold=(ci==1),size=10)
            c.alignment=lft() if ci in (1,8) else ctr()
            c.border=bd()

map_rows = []
for p in potential_hosts:
    latlon = geocode_address(p["addr"], cfg)
    tags = []
    if p["can_drinks"] and p["name"] not in hosted_drinks:
        tags.append("Apéros non desservis")
    if p["name"] in hosted_drinks:
        tags.append("Apéros réels")
    if p["can_dinner"] and p["name"] not in hosted_dinner:
        tags.append("Dîners non desservis")
    if p["name"] in hosted_dinner:
        tags.append("Dîners réels")
    map_rows.append({
        "label": p["name"],
        "addr": p["addr"],
        "lat": (latlon[0] if latlon else None),
        "lon": (latlon[1] if latlon else None),
        "tags": " | ".join(tags),
        "can_drinks": p["can_drinks"],
        "can_dinner": p["can_dinner"],
        "actual_drinks": (p["name"] in hosted_drinks),
        "actual_dinner": (p["name"] in hosted_dinner),
    })

map_rows.sort(key=lambda x: x["label"].casefold())

# Export a real KML source file for Google My Maps
map_kml_path = f"{os.path.splitext(out)[0]}_hotes_potentiels_mymaps.kml"
with open(map_kml_path, "w", encoding="utf-8") as kf:
    kf.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    kf.write('<kml xmlns="http://www.opengis.net/kml/2.2">\n')
    kf.write('  <Document>\n')
    kf.write(f'    <name>{xml_esc(event_title)}: apéros - dîners</name>\n')

    write_had_styles(kf)

    categories = [
        ("Apéros non desservis", "icon-1517-BDBDBD", lambda m: m["can_drinks"] and not m["actual_drinks"], "Apero potentiel non desservi"),
        ("Apéros réels", "icon-1517-0288D1", lambda m: m["actual_drinks"], "Apero reel"),
        ("Dîners non desservis", "icon-1577-BDBDBD", lambda m: m["can_dinner"] and not m["actual_dinner"], "Diner potentiel non desservi"),
        ("Dîners réels", "icon-1577-FF5252", lambda m: m["actual_dinner"], "Diner reel"),
    ]
    for category_label, style_id, pred, desc_label in categories:
        kf.write('    <Folder>\n')
        kf.write(f'      <name>{xml_esc(category_label)}</name>\n')
        for m in map_rows:
            if pred(m):
                write_kml_placemark(
                    kf,
                    name=m["label"],
                    address=m["addr"],
                    style_id=style_id,
                    category_label=desc_label,
                    lat=m["lat"],
                    lon=m["lon"],
                )
        kf.write('    </Folder>\n')

    dessert_latlon = geocode_address(dessert_addr, cfg)
    kf.write('    <Folder>\n')
    kf.write('      <name>Dessert</name>\n')
    write_kml_placemark(
        kf,
        name=cfg.get("dessert_address", "Dessert"),
        address=dessert_addr,
        style_id="icon-1762-7CB342-nodesc",
        category_label="Dessert",
        lat=(dessert_latlon[0] if dessert_latlon else None),
        lon=(dessert_latlon[1] if dessert_latlon else None),
    )
    kf.write('    </Folder>\n')

    kf.write('  </Document>\n')
    kf.write('</kml>\n')

participants_kml_path = f"{os.path.splitext(out)[0]}_participants_mymaps.kml"
participant_rows = []
for p in sorted(people_csv, key=lambda x: x["name"].casefold()):
    latlon = geocode_address(p["addr"], cfg)
    participant_rows.append({
        "label": p["name"],
        "addr": p["addr"],
        "lat": (latlon[0] if latlon else None),
        "lon": (latlon[1] if latlon else None),
    })

with open(participants_kml_path, "w", encoding="utf-8") as kf:
    kf.write('<?xml version="1.0" encoding="UTF-8"?>\n')
    kf.write('<kml xmlns="http://www.opengis.net/kml/2.2">\n')
    kf.write('  <Document>\n')
    kf.write(f'    <name>{xml_esc(event_title)}: participants</name>\n')

    write_had_styles(kf)

    kf.write('    <Folder>\n')
    kf.write('      <name>Participants</name>\n')
    for p in participant_rows:
        write_kml_placemark(
            kf,
            name=p["label"],
            address=p["addr"],
            style_id="icon-1603-0288D1",
            category_label="Participant",
            lat=p["lat"],
            lon=p["lon"],
        )
    kf.write('    </Folder>\n')
    kf.write('  </Document>\n')
    kf.write('</kml>\n')

# ═══ SHEET 5 — Stats ═════════════════════════════════════════════════════════
ws5=wb.create_sheet("Statistiques")
ws5.sheet_view.showGridLines=False
for col,w in zip([1,2,3],[36,22,30]): cw(ws5,col,w)
ws5.merge_cells("A1:C1")
ws5["A1"]="📈 STATISTIQUES"
ws5["A1"].font=Font(name='Arial',bold=True,size=14,color='FFFFFF')
ws5["A1"].fill=fill("263238"); ws5["A1"].alignment=ctr(); ws5.row_dimensions[1].height=30

total_walk=sum(r['total'] for r in rows)
max_r=max(rows,key=lambda r:r['total']); min_r=min(rows,key=lambda r:r['total'])
row_by_name = {r['name']: r for r in rows}
dinner_host_walks = []
for host_name in sorted(dinner_groups.keys()):
    host_row = row_by_name.get(host_name)
    if host_row:
        dinner_host_walks.append({
            'name': host_name,
            'w2': host_row['w2'],
            'drinks_host': host_row['drinks_host'],
        })

stats=[
    ("","",""),("👥 PARTICIPANTS","",""),
    ("Nombre total",len(rows),"personnes"),
    ("","",""),("🍸 APERITIF","",""),("Hotes",len(drinks_groups),""),
]
for h,g in drinks_groups.items():
    stats.append((f"  Chez {h}",f"{len(g)} participants",""))
stats+=[("","",""),("🍽️ DINER","",""),("Hotes",len(dinner_groups),"")]
for h,g in dinner_groups.items():
    stats.append((f"  Chez {h}",f"{len(g)} participants",""))
stats += [("","",""),("🏠 HOTES DINER — Apero→Diner","","")]
if dinner_host_walks:
    avg_host_walk = sum(r['w2'] for r in dinner_host_walks) / len(dinner_host_walks)
    max_host_walk = max(dinner_host_walks, key=lambda r: r['w2'])
    for r in dinner_host_walks:
        stats.append((f"  {r['name']}",f"{r['w2']:.1f} min",f"depuis chez {r['drinks_host']}"))
    stats.append(("  Moyenne hotes",f"{avg_host_walk:.1f} min",""))
    stats.append(("  Maximum hote",f"{max_host_walk['w2']:.1f} min",f"({max_host_walk['name']})"))
else:
    stats.append(("  (aucun hote diner)","",""))
stats+=[
    ("","",""),("🚶 TEMPS DE MARCHE","",""),
    ("Moyenne / personne",f"{total_walk/len(rows):.1f} min",""),
    ("Maximum",f"{max_r['total']:.1f} min",f"({max_r['name']})"),
    ("Minimum",f"{min_r['total']:.1f} min",f"({min_r['name']})"),
]
sections={"👥 PARTICIPANTS","🍸 APERITIF","🍽️ DINER","🏠 HOTES DINER — Apero→Diner","🚶 TEMPS DE MARCHE"}
for ri,(a,b,c) in enumerate(stats,2):
    ws5.row_dimensions[ri].height=20
    for ci,v in enumerate([a,b,c],1):
        cell=ws5.cell(row=ri,column=ci,value=v); cell.border=bd()
        if str(a) in sections:
            cell.font=Font(name='Arial',bold=True,size=11,color='FFFFFF')
            cell.fill=fill("455A64"); cell.alignment=lft()
        elif str(a).startswith("  "):
            cell.font=cfont(size=10); cell.fill=fill("ECEFF1"); cell.alignment=lft()
        else:
            cell.font=cfont(bold=(ci==1),size=10)
            cell.fill=fill("F5F5F5" if ri%2==0 else "FFFFFF"); cell.alignment=lft() if ci==1 else ctr()

# Auto-size the last stats column to fit labels like "depuis chez ...".
max_col3_len = max((len(str(c or "")) for _, _, c in stats), default=0)
cw(ws5, 3, max(30, min(60, max_col3_len + 4)))

out_dir = os.path.dirname(out)
if out_dir:
    os.makedirs(out_dir, exist_ok=True)
geo_cache_dir = os.path.dirname(geocode_cache_path)
if geo_cache_dir:
    os.makedirs(geo_cache_dir, exist_ok=True)
with open(geocode_cache_path, "w") as gf:
    json.dump(geocode_cache, gf, ensure_ascii=False, indent=2)
wb.save(out)
print(f"Excel saved: {out}")
print(f"My Maps KML saved: {map_kml_path}")
print(f"Participants KML saved: {participants_kml_path}")
